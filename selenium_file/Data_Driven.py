import time

from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By


def test_view11():
    workbook = load_workbook("test_data.xlsx")

    sheet = workbook.active

    driver = webdriver.Firefox()

    driver.maximize_window()

    for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
        username = row[0]
        password = row[1]
        driver.get("https://www.saucedemo.com/v1/")
        time.sleep(2)
        username_field = driver.find_element(By.ID, value="user-name")
        password_filed = driver.find_element(By.ID, value="password")
        username_field.send_keys(username)
        password_filed.send_keys(password)
        login_filed = driver.find_element(By.ID, value="login-button")
        assert not login_filed.get_attribute("disable")
        login_filed.click()